from __future__ import annotations

import re
import json
from io import BytesIO
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any, Dict, List, Optional, Tuple
from datetime import datetime, date

import fitz  # PyMuPDF
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill

# =========================================================
# INVOICE IQ AGENT - HACKATHON ENERGY INVOICE DEMO
# =========================================================
APP_TITLE = "Invoice IQ Agent — Foundry Polished Demo"
TOLERANCE = 0.05
AMOUNT_PATTERN = r"(?:-?[0-9,]+(?:\.\d+)?|\([0-9,]+(?:\.\d+)?\))"

# Exact columns requested, matching your original electricity-invoice output order.
OUTPUT_COLUMNS = [
    "Provider",
    "Production Month",
    "From",
    "To",
    "Invoice Date",
    "Power Factor",
    "Load Factor",
    "Actual Demand (KW)",
    "Billing Demand (KW)",
    "4CP Charges Qty (KW)",
    "4CP Charges Rate ($/KW)",
    "4CP Charges ($)",
    "Usage - Actual KWH",
    "UOM",
    "Energy Charge",
    "Nodal Congestion Charge",
    "Market Securitization (Debt) Financing - Default Charge",
    "Prior Period Pass Through Charge",
    "ERCOT Cont Reserve Serv (ECRS)",
    "Firm Fuel Supply Service",
    "Firm Fuel Supply Service - Backbill",
    "Market Securitization - Uplift Charge",
    "TX-ERCOT Admin Fees - CIL",
    "Transmission Charges",
    "Taxes & PUC Assessment Charge",
    "Ancilliary Service Obligation Adjustment",
    "Other Taxes",
    "Bill Total",
]

DATE_COLUMNS = ["Production Month", "From", "To", "Invoice Date"]
PERCENT_COLUMNS = ["Power Factor", "Load Factor"]
NUMERIC_COLUMNS = [
    "Actual Demand (KW)",
    "Billing Demand (KW)",
    "4CP Charges Qty (KW)",
    "4CP Charges Rate ($/KW)",
    "4CP Charges ($)",
    "Usage - Actual KWH",
    "Energy Charge",
    "Nodal Congestion Charge",
    "Market Securitization (Debt) Financing - Default Charge",
    "Prior Period Pass Through Charge",
    "ERCOT Cont Reserve Serv (ECRS)",
    "Firm Fuel Supply Service",
    "Firm Fuel Supply Service - Backbill",
    "Market Securitization - Uplift Charge",
    "TX-ERCOT Admin Fees - CIL",
    "Transmission Charges",
    "Taxes & PUC Assessment Charge",
    "Ancilliary Service Obligation Adjustment",
    "Other Taxes",
    "Bill Total",
]

# These components are summed for the demo-level Bill Total validation.
# 4CP Charges are validated separately as qty * rate.
BILL_COMPONENT_COLUMNS = [
    "Energy Charge",
    "Nodal Congestion Charge",
    "Market Securitization (Debt) Financing - Default Charge",
    "Prior Period Pass Through Charge",
    "ERCOT Cont Reserve Serv (ECRS)",
    "Firm Fuel Supply Service",
    "Firm Fuel Supply Service - Backbill",
    "Market Securitization - Uplift Charge",
    "TX-ERCOT Admin Fees - CIL",
    "Transmission Charges",
    "Taxes & PUC Assessment Charge",
    "Ancilliary Service Obligation Adjustment",
    "Other Taxes",
]

REQUIRED_COLUMNS = [
    "Provider",
    "Production Month",
    "From",
    "To",
    "Invoice Date",
    "Usage - Actual KWH",
    "Energy Charge",
    "Transmission Charges",
    "Taxes & PUC Assessment Charge",
    "Bill Total",
]

LINE_ITEM_COLUMNS = [
    "Source File",
    "Provider",
    "Invoice Date",
    "Description",
    "Qty",
    "Rate",
    "Amount",
]

FLAG_COLUMNS = [
    "Source File",
    "Rule",
    "Severity",
    "Status",
    "Message",
    "Suggested Review",
]

VALIDATION_COLUMNS = [
    "Source File",
    "Review Status",
    "Confidence",
    "Required Fields",
    "Bill Total Math",
    "4CP Math",
    "Line Item Sum",
    "Validation Notes",
]

# =========================================================
# HELPERS
# =========================================================
def normalize_text(text: str) -> str:
    text = text.replace("\u00a0", " ")
    text = text.replace("–", "-").replace("—", "-")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{2,}", "\n", text)
    return text.strip()


def to_float(value: Any) -> Optional[float]:
    if value is None:
        return None

    s = str(value).strip()
    s = s.replace("$", "").replace(",", "")
    s = s.replace("(", "-").replace(")", "")
    s = s.replace("KWH", "").replace("kWh", "")
    s = s.replace("KW", "").replace("kW", "")
    s = s.replace("/KWH", "").replace("/kWh", "")
    s = s.replace("/KW", "").replace("/kW", "")
    s = s.strip()

    if not s:
        return None

    try:
        return float(s)
    except ValueError:
        return None


def percent_text_to_decimal(value: Any) -> Optional[float]:
    if value is None:
        return None
    s = str(value).strip().replace("%", "")
    num = to_float(s)
    if num is None:
        return None
    return num / 100.0 if num > 1 else num


def parse_date(value: Any) -> Optional[date]:
    if value is None:
        return None

    s = str(value).strip()
    if not s:
        return None

    formats = [
        "%m/%d/%Y",
        "%m/%d/%y",
        "%Y-%m-%d",
        "%b %d, %Y",
        "%B %d, %Y",
        "%B-%y",
        "%b-%y",
    ]

    for fmt in formats:
        try:
            dt = datetime.strptime(s, fmt).date()
            if fmt in {"%B-%y", "%b-%y"}:
                return date(dt.year, dt.month, 1)
            return dt
        except ValueError:
            pass

    return None


def nearly_equal(a: Optional[float], b: Optional[float], tol: float = TOLERANCE) -> bool:
    if a is None or b is None:
        return False
    return abs(float(a) - float(b)) <= tol


def extract_text_pymupdf(pdf_path: Path) -> str:
    doc = fitz.open(str(pdf_path))
    parts: List[str] = []
    for page in doc:
        parts.append(page.get_text("text") or "")
    doc.close()
    return normalize_text("\n".join(parts))


def find_labeled_value(text: str, label: str) -> Optional[str]:
    """Extract values from lines like: Label: value"""
    label_lower = label.lower()
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if line.lower().startswith(label_lower + ":"):
            return line.split(":", 1)[1].strip()
    return None


def clean_display_df(df: pd.DataFrame) -> pd.DataFrame:
    """Show dates as readable strings in Streamlit while keeping export formats intact."""
    out = df.copy()
    for col in DATE_COLUMNS:
        if col in out.columns:
            out[col] = pd.to_datetime(out[col], errors="coerce").dt.strftime("%m/%d/%Y")
    for col in PERCENT_COLUMNS:
        if col in out.columns:
            out[col] = pd.to_numeric(out[col], errors="coerce").map(lambda x: f"{x:.1%}" if pd.notna(x) else "")
    return out

# =========================================================
# EXTRACTION
# =========================================================
def extract_invoice_fields(text: str, source_file: str) -> Dict[str, Any]:
    row: Dict[str, Any] = {col: None for col in OUTPUT_COLUMNS}

    for col in OUTPUT_COLUMNS:
        value = find_labeled_value(text, col)
        if value is not None:
            row[col] = value

    # Data type conversion
    for col in DATE_COLUMNS:
        row[col] = parse_date(row.get(col))

    for col in PERCENT_COLUMNS:
        row[col] = percent_text_to_decimal(row.get(col))

    for col in NUMERIC_COLUMNS:
        row[col] = to_float(row.get(col))

    # Defaults used by the demo if the source invoice leaves these blank.
    if not row.get("Provider"):
        row["Provider"] = "Texas GLO/State Power Program"
    if not row.get("UOM"):
        row["UOM"] = "KWH"

    row["Source File"] = source_file
    return row


def extract_line_items(text: str, row: Dict[str, Any], source_file: str) -> List[Dict[str, Any]]:
    items: List[Dict[str, Any]] = []

    # Expected synthetic format:
    # Line Item: Energy Charge | Qty: 2350000.00 | Rate: 0.050000 | Amount: 117500.00
    pattern = re.compile(
        rf"^Line Item:\s*(?P<description>.*?)\s*\|\s*"
        rf"Qty:\s*(?P<qty>{AMOUNT_PATTERN})\s*\|\s*"
        rf"Rate:\s*\$?(?P<rate>{AMOUNT_PATTERN})\s*\|\s*"
        rf"Amount:\s*\$?(?P<amount>{AMOUNT_PATTERN})\s*$",
        flags=re.I,
    )

    for raw_line in text.splitlines():
        line = raw_line.strip()
        m = pattern.match(line)
        if not m:
            continue

        items.append(
            {
                "Source File": source_file,
                "Provider": row.get("Provider"),
                "Invoice Date": row.get("Invoice Date"),
                "Description": m.group("description").strip(),
                "Qty": to_float(m.group("qty")),
                "Rate": to_float(m.group("rate")),
                "Amount": to_float(m.group("amount")),
            }
        )

    return items

# =========================================================
# VALIDATION / FLAGS
# =========================================================
def add_flag(
    flags: List[Dict[str, Any]],
    source_file: str,
    rule: str,
    severity: str,
    status: str,
    message: str,
    suggested_review: str,
) -> None:
    flags.append(
        {
            "Source File": source_file,
            "Rule": rule,
            "Severity": severity,
            "Status": status,
            "Message": message,
            "Suggested Review": suggested_review,
        }
    )


def validate_invoice(row: Dict[str, Any], line_items: List[Dict[str, Any]]) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    source_file = row.get("Source File", "Unknown")
    flags: List[Dict[str, Any]] = []
    notes: List[str] = []

    # Required fields
    missing = [col for col in REQUIRED_COLUMNS if row.get(col) in [None, ""] or pd.isna(row.get(col))]
    if missing:
        required_status = "FAIL"
        msg = "Missing required field(s): " + ", ".join(missing)
        notes.append(msg)
        add_flag(flags, source_file, "Required fields", "Critical", "FAIL", msg, "Review the source PDF labels or parser mapping.")
    else:
        required_status = "PASS"

    # 4CP qty * rate = amount
    cp4_qty = row.get("4CP Charges Qty (KW)")
    cp4_rate = row.get("4CP Charges Rate ($/KW)")
    cp4_amount = row.get("4CP Charges ($)")
    if cp4_qty is not None and cp4_rate is not None and cp4_amount is not None:
        expected_4cp = round(float(cp4_qty) * float(cp4_rate), 2)
        if nearly_equal(expected_4cp, float(cp4_amount)):
            cp4_status = "PASS"
        else:
            cp4_status = "FAIL"
            msg = f"4CP mismatch: qty * rate = ${expected_4cp:,.2f}, extracted 4CP amount = ${float(cp4_amount):,.2f}."
            notes.append(msg)
            add_flag(flags, source_file, "4CP Charges Qty * Rate", "Critical", "FAIL", msg, "Check 4CP quantity, rate, and amount.")
    else:
        cp4_status = "WARN"
        msg = "Cannot validate 4CP math because qty, rate, or amount is missing."
        notes.append(msg)
        add_flag(flags, source_file, "4CP Charges Qty * Rate", "Warning", "WARN", msg, "Review 4CP fields.")

    # Sum bill components = Bill Total
    component_values = [float(row.get(col) or 0) for col in BILL_COMPONENT_COLUMNS]
    expected_total = round(sum(component_values), 2)
    bill_total = row.get("Bill Total")
    if bill_total is not None:
        if nearly_equal(expected_total, float(bill_total)):
            total_status = "PASS"
        else:
            total_status = "FAIL"
            msg = f"Bill total mismatch: components sum to ${expected_total:,.2f}, extracted Bill Total = ${float(bill_total):,.2f}."
            notes.append(msg)
            add_flag(flags, source_file, "Bill components sum = Bill Total", "Critical", "FAIL", msg, "Review all charge components and final total.")
    else:
        total_status = "WARN"
        msg = "Cannot validate Bill Total because Bill Total is missing."
        notes.append(msg)
        add_flag(flags, source_file, "Bill components sum = Bill Total", "Warning", "WARN", msg, "Review Bill Total extraction.")

    # Line item sum = Bill Total for the synthetic demo
    if line_items and bill_total is not None:
        line_sum = round(sum(float(item.get("Amount") or 0) for item in line_items), 2)
        if nearly_equal(line_sum, float(bill_total)):
            line_status = "PASS"
        else:
            line_status = "FAIL"
            msg = f"Line item mismatch: line items sum to ${line_sum:,.2f}, extracted Bill Total = ${float(bill_total):,.2f}."
            notes.append(msg)
            add_flag(flags, source_file, "Line items sum = Bill Total", "Critical", "FAIL", msg, "Review extracted line items.")
    else:
        line_status = "WARN"
        msg = "No line items were extracted or Bill Total is missing."
        notes.append(msg)
        add_flag(flags, source_file, "Line item extraction", "Warning", "WARN", msg, "Use the included synthetic invoice format or improve the table parser.")

    critical_count = sum(1 for f in flags if f["Severity"] == "Critical")
    review_status = "PASS" if critical_count == 0 else "REVIEW"

    found_required = len(REQUIRED_COLUMNS) - len(missing)
    field_score = found_required / len(REQUIRED_COLUMNS)
    validation_score = sum(x == "PASS" for x in [required_status, cp4_status, total_status, line_status]) / 4
    line_score = 1.0 if line_items else 0.0
    confidence = round(field_score * 0.50 + validation_score * 0.40 + line_score * 0.10, 2)

    validation = {
        "Source File": source_file,
        "Review Status": review_status,
        "Confidence": confidence,
        "Required Fields": required_status,
        "Bill Total Math": total_status,
        "4CP Math": cp4_status,
        "Line Item Sum": line_status,
        "Validation Notes": " | ".join(notes) if notes else "No exceptions",
    }

    return validation, flags

# =========================================================
# AGENT Q&A
# =========================================================
def ask_agent_rule_based(question: str, summary_df: pd.DataFrame, line_items_df: pd.DataFrame, flags_df: pd.DataFrame, validation_df: pd.DataFrame) -> str:
    if summary_df.empty:
        return "No invoice has been processed yet."

    q = question.lower().strip()

    if any(term in q for term in ["how many", "count", "invoice"]):
        return f"I processed {len(summary_df)} invoice(s). {len(validation_df[validation_df['Review Status'] == 'REVIEW'])} need review."

    if any(term in q for term in ["column", "field", "extract", "output"]):
        return "The Summary tab uses the requested electricity-invoice columns: " + ", ".join(OUTPUT_COLUMNS) + "."

    if any(term in q for term in ["total", "match", "math", "bill"]):
        fails = validation_df[validation_df["Bill Total Math"] != "PASS"]
        if fails.empty:
            return "Yes. For every processed invoice, the bill components match the extracted Bill Total within the demo tolerance."
        return "Some invoices failed total validation: " + ", ".join(fails["Source File"].astype(str).tolist())

    if any(term in q for term in ["4cp", "qty", "rate"]):
        fails = validation_df[validation_df["4CP Math"] != "PASS"]
        if fails.empty:
            return "Yes. 4CP Qty x 4CP Rate matches 4CP Charges for every processed invoice."
        return "Some invoices failed 4CP validation: " + ", ".join(fails["Source File"].astype(str).tolist())

    if any(term in q for term in ["flag", "exception", "issue", "wrong", "review"]):
        if flags_df.empty:
            return "No flags were created. All processed invoices passed the current demo validation rules."
        return "Flag summary: " + " | ".join(flags_df["Message"].astype(str).head(8).tolist())

    if any(term in q for term in ["highest", "largest", "most expensive"]):
        idx = pd.to_numeric(summary_df["Bill Total"], errors="coerce").idxmax()
        row = summary_df.loc[idx]
        return f"The largest Bill Total is ${row['Bill Total']:,.2f} from source file {validation_df.loc[idx, 'Source File']}."

    if any(term in q for term in ["average", "rate", "kwh"]):
        total_bill = pd.to_numeric(summary_df["Bill Total"], errors="coerce").sum()
        total_kwh = pd.to_numeric(summary_df["Usage - Actual KWH"], errors="coerce").sum()
        if total_kwh == 0:
            return "I cannot calculate average $/KWH because usage is zero or missing."
        return f"Across the processed invoices, average cost is approximately ${total_bill / total_kwh:.6f} per KWH."

    if any(term in q for term in ["line", "item", "items"]):
        return f"I extracted {len(line_items_df)} line item rows across {len(summary_df)} invoice(s)."

    return "Try asking: How many invoices were processed? Do totals match? Which invoices were flagged? What columns were extracted? What is the average $/KWH?"



def _compact_dataframe(df: pd.DataFrame, max_rows: int = 20) -> str:
    """Return compact JSON records for agent grounding without overwhelming the prompt."""
    if df.empty:
        return "[]"
    safe_df = df.head(max_rows).copy()
    for col in safe_df.columns:
        if pd.api.types.is_datetime64_any_dtype(safe_df[col]):
            safe_df[col] = safe_df[col].dt.strftime("%m/%d/%Y")
    return safe_df.to_json(orient="records", date_format="iso")


def build_agent_context(summary_df: pd.DataFrame, line_items_df: pd.DataFrame, flags_df: pd.DataFrame, validation_df: pd.DataFrame) -> str:
    """Build invoice context to send to the Microsoft Foundry agent."""
    metrics = {
        "invoice_count": int(len(summary_df)),
        "review_count": int((validation_df["Review Status"] == "REVIEW").sum()) if not validation_df.empty else 0,
        "pass_count": int((validation_df["Review Status"] == "PASS").sum()) if not validation_df.empty else 0,
        "flag_count": int(len(flags_df)),
        "total_bill": float(pd.to_numeric(summary_df.get("Bill Total"), errors="coerce").sum()) if "Bill Total" in summary_df else 0,
        "total_kwh": float(pd.to_numeric(summary_df.get("Usage - Actual KWH"), errors="coerce").sum()) if "Usage - Actual KWH" in summary_df else 0,
    }

    context = {
        "purpose": "Invoice IQ Agent dashboard context from uploaded synthetic energy invoices.",
        "metrics": metrics,
        "summary_rows": json.loads(_compact_dataframe(summary_df[OUTPUT_COLUMNS + ["Source File"]] if "Source File" in summary_df.columns else summary_df)),
        "validation_rows": json.loads(_compact_dataframe(validation_df)),
        "flag_rows": json.loads(_compact_dataframe(flags_df)),
        "line_item_sample": json.loads(_compact_dataframe(line_items_df, max_rows=15)),
        "status_rules": {
            "PASS": "No critical validation exceptions.",
            "REVIEW": "At least one validation issue needs analyst review.",
            "FAIL": "Used inside specific validation checks, not the overall dashboard review label."
        }
    }
    return json.dumps(context, indent=2)


def ask_agent(question: str, summary_df: pd.DataFrame, line_items_df: pd.DataFrame, flags_df: pd.DataFrame, validation_df: pd.DataFrame) -> str:
    """Use Microsoft Foundry agent when configured; fall back to local deterministic Q&A."""
    if summary_df.empty:
        return "No invoice has been processed yet."

    try:
        from agent_client import InvoiceIQAgentClient
        client = InvoiceIQAgentClient()
        prompt = f"""
You are answering inside the Invoice IQ Streamlit demo.
Use the invoice dashboard context below. Be concise, business-facing, and specific.
Do not invent values. If a value is not in the context, say it is not available.

User question:
{question}

Invoice dashboard context:
{build_agent_context(summary_df, line_items_df, flags_df, validation_df)}
""".strip()
        return client.send_message(prompt, use_history=False)
    except Exception as exc:
        fallback = ask_agent_rule_based(question, summary_df, line_items_df, flags_df, validation_df)
        return f"{fallback}\n\n_Foundry agent fallback used: {exc}_"

# =========================================================
# WORKBOOK EXPORT
# =========================================================
def build_excel_workbook(summary_df: pd.DataFrame, line_items_df: pd.DataFrame, flags_df: pd.DataFrame, validation_df: pd.DataFrame) -> bytes:
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl", datetime_format="m/d/yyyy", date_format="m/d/yyyy") as writer:
        # Summary uses the exact requested columns only.
        summary_df[OUTPUT_COLUMNS].to_excel(writer, sheet_name="Summary", index=False)
        line_items_df.to_excel(writer, sheet_name="Line Items", index=False)
        flags_df.to_excel(writer, sheet_name="Flags_Exceptions", index=False)
        validation_df.to_excel(writer, sheet_name="Validation", index=False)

        wb = writer.book

        money_columns = set(BILL_COMPONENT_COLUMNS + ["4CP Charges ($)", "Bill Total", "Amount"])
        rate_columns = {"4CP Charges Rate ($/KW)", "Rate"}
        number_columns = {"Actual Demand (KW)", "Billing Demand (KW)", "4CP Charges Qty (KW)", "Usage - Actual KWH", "Qty"}
        percent_columns = set(PERCENT_COLUMNS)
        date_columns = {"Production Month", "From", "To", "Invoice Date"}

        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.fill = PatternFill("solid", fgColor="1F4E78")

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    header = ws.cell(row=1, column=cell.column).value
                    if cell.value is None:
                        continue
                    if header in money_columns:
                        cell.number_format = "$#,##0.00"
                    elif header in rate_columns:
                        cell.number_format = "$0.000000"
                    elif header in number_columns:
                        cell.number_format = "#,##0.00"
                    elif header in percent_columns:
                        cell.number_format = "0.0%"
                    elif header in date_columns:
                        if header == "Production Month":
                            cell.number_format = "mmmm-yy"
                        else:
                            cell.number_format = "m/d/yyyy"
                    elif header == "Confidence":
                        cell.number_format = "0.00"

            for col_cells in ws.columns:
                max_len = 0
                col_letter = col_cells[0].column_letter
                for c in col_cells:
                    val = "" if c.value is None else str(c.value)
                    max_len = max(max_len, max(len(part) for part in val.split("\n")))
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 38)

            ws.row_dimensions[1].height = 42

    return output.getvalue()

# =========================================================
# PROCESSOR
# =========================================================
def process_pdf(pdf_path: Path, display_name: str) -> Tuple[Dict[str, Any], List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any], str]:
    text = extract_text_pymupdf(pdf_path)
    if not text.strip():
        raise ValueError("No readable text found. Use text-based sample PDFs for the hackathon demo.")

    row = extract_invoice_fields(text, display_name)
    line_items = extract_line_items(text, row, display_name)
    validation, flags = validate_invoice(row, line_items)
    return row, line_items, flags, validation, text


def process_uploaded_files(uploaded_files: List[Any]) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, str]:
    rows: List[Dict[str, Any]] = []
    all_line_items: List[Dict[str, Any]] = []
    all_flags: List[Dict[str, Any]] = []
    validations: List[Dict[str, Any]] = []
    raw_text_parts: List[str] = []

    for uploaded_file in uploaded_files:
        with NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            tmp.write(uploaded_file.getbuffer())
            pdf_path = Path(tmp.name)

        try:
            row, line_items, flags, validation, raw_text = process_pdf(pdf_path, uploaded_file.name)
            rows.append(row)
            all_line_items.extend(line_items)
            all_flags.extend(flags)
            validations.append(validation)
            raw_text_parts.append(f"===== {uploaded_file.name} =====\n{raw_text}")
        except Exception as exc:
            rows.append({**{col: None for col in OUTPUT_COLUMNS}, "Provider": "ERROR", "Source File": uploaded_file.name})
            validation = {
                "Source File": uploaded_file.name,
                "Review Status": "REVIEW",
                "Confidence": 0.0,
                "Required Fields": "FAIL",
                "Bill Total Math": "FAIL",
                "4CP Math": "FAIL",
                "Line Item Sum": "FAIL",
                "Validation Notes": f"ERROR: {exc}",
            }
            validations.append(validation)
            all_flags.append(
                {
                    "Source File": uploaded_file.name,
                    "Rule": "Processing",
                    "Severity": "Critical",
                    "Status": "FAIL",
                    "Message": f"Could not process PDF: {exc}",
                    "Suggested Review": "Confirm the PDF is text-based and follows the demo format.",
                }
            )

    summary_df = pd.DataFrame(rows, columns=[*OUTPUT_COLUMNS, "Source File"])
    line_items_df = pd.DataFrame(all_line_items, columns=LINE_ITEM_COLUMNS)
    flags_df = pd.DataFrame(all_flags, columns=FLAG_COLUMNS)
    validation_df = pd.DataFrame(validations, columns=VALIDATION_COLUMNS)

    for col in DATE_COLUMNS:
        if col in summary_df.columns:
            summary_df[col] = pd.to_datetime(summary_df[col], errors="coerce")
        if col in line_items_df.columns:
            line_items_df[col] = pd.to_datetime(line_items_df[col], errors="coerce")

    for col in PERCENT_COLUMNS + NUMERIC_COLUMNS:
        if col in summary_df.columns:
            summary_df[col] = pd.to_numeric(summary_df[col], errors="coerce")

    for col in ["Qty", "Rate", "Amount"]:
        if col in line_items_df.columns:
            line_items_df[col] = pd.to_numeric(line_items_df[col], errors="coerce")

    return summary_df, line_items_df, flags_df, validation_df, "\n\n".join(raw_text_parts)

# =========================================================
# STREAMLIT UI - PROFESSIONAL HACKATHON DEMO
# =========================================================
def inject_professional_css() -> None:
    st.markdown(
        """
        <style>
            :root {
                --iq-bg: #f6f8fb;
                --iq-card: #ffffff;
                --iq-text: #111827;
                --iq-muted: #6b7280;
                --iq-blue: #2563eb;
                --iq-green: #16834a;
                --iq-gold: #d99a16;
                --iq-red: #dc2626;
                --iq-border: #e5e7eb;
            }

            .stApp {
                background: radial-gradient(circle at top left, #eaf1ff 0, #f8fafc 28%, #ffffff 100%);
            }

            section[data-testid="stSidebar"] {
                background: #0f172a;
            }

            section[data-testid="stSidebar"] * {
                color: #e5e7eb !important;
            }

            .block-container {
                padding-top: 1.4rem;
                padding-bottom: 3rem;
                max-width: 1320px;
            }

            .iq-hero {
                border: 1px solid rgba(37, 99, 235, 0.16);
                background: linear-gradient(135deg, rgba(37,99,235,.12), rgba(22,131,74,.08) 48%, rgba(255,255,255,.96));
                border-radius: 26px;
                padding: 30px 34px;
                box-shadow: 0 18px 45px rgba(15, 23, 42, 0.08);
                margin-bottom: 22px;
                position: relative;
                overflow: hidden;
            }

            .iq-hero:after {
                content: "";
                width: 210px;
                height: 210px;
                border-radius: 999px;
                background: rgba(37,99,235,0.10);
                position: absolute;
                right: -55px;
                top: -80px;
            }

            .iq-eyebrow {
                display: inline-flex;
                gap: 8px;
                align-items: center;
                padding: 7px 12px;
                border-radius: 999px;
                background: rgba(255,255,255,.78);
                border: 1px solid rgba(37,99,235,.18);
                font-size: 0.82rem;
                font-weight: 700;
                color: #1d4ed8;
                letter-spacing: .02em;
                text-transform: uppercase;
            }

            .iq-title {
                margin-top: 18px;
                font-size: 3.15rem;
                line-height: 1.02;
                font-weight: 900;
                color: var(--iq-text);
                letter-spacing: -0.045em;
            }

            .iq-subtitle {
                margin-top: 12px;
                font-size: 1.18rem;
                color: #475569;
                max-width: 850px;
            }

            .iq-chip-row {
                margin-top: 22px;
                display: flex;
                flex-wrap: wrap;
                gap: 10px;
            }

            .iq-chip {
                display: inline-flex;
                align-items: center;
                gap: 8px;
                border-radius: 999px;
                padding: 9px 13px;
                background: #ffffff;
                border: 1px solid #dbe3ef;
                box-shadow: 0 5px 18px rgba(15, 23, 42, 0.05);
                font-weight: 700;
                color: #1f2937;
                font-size: .92rem;
            }

            .iq-card {
                background: rgba(255,255,255,.92);
                border: 1px solid var(--iq-border);
                border-radius: 22px;
                padding: 22px;
                box-shadow: 0 14px 35px rgba(15, 23, 42, 0.06);
                height: 100%;
            }

            .iq-card h3, .iq-card h4 {
                margin: 0 0 8px 0;
                color: var(--iq-text);
                letter-spacing: -0.02em;
            }

            .iq-card p {
                color: var(--iq-muted);
                margin: 0;
                font-size: .95rem;
            }

            .iq-kpi {
                background: #ffffff;
                border: 1px solid #e3e9f2;
                border-radius: 20px;
                padding: 18px 20px;
                box-shadow: 0 12px 28px rgba(15, 23, 42, .06);
                min-height: 128px;
            }

            .iq-kpi-label {
                font-size: .83rem;
                color: #64748b;
                font-weight: 800;
                text-transform: uppercase;
                letter-spacing: .06em;
            }

            .iq-kpi-value {
                margin-top: 10px;
                font-size: 2rem;
                line-height: 1;
                font-weight: 900;
                letter-spacing: -0.04em;
                color: #0f172a;
            }

            .iq-kpi-caption {
                margin-top: 10px;
                font-size: .9rem;
                color: #64748b;
            }

            .iq-flow {
                background: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 18px;
                padding: 18px;
                box-shadow: 0 10px 25px rgba(15, 23, 42, .05);
                min-height: 150px;
            }

            .iq-step-number {
                width: 36px;
                height: 36px;
                border-radius: 999px;
                display: inline-flex;
                align-items: center;
                justify-content: center;
                background: #2563eb;
                color: white;
                font-weight: 900;
                margin-bottom: 14px;
            }

            .iq-flow h4 {
                margin: 0 0 6px 0;
                font-size: 1.05rem;
                color: #111827;
            }

            .iq-flow p {
                margin: 0;
                color: #64748b;
                font-size: .9rem;
            }

            .iq-status-pass {
                background: #ecfdf5;
                color: #047857;
                border: 1px solid #a7f3d0;
                padding: 5px 10px;
                border-radius: 999px;
                font-weight: 900;
            }

            .iq-status-review {
                background: #fff7ed;
                color: #c2410c;
                border: 1px solid #fed7aa;
                padding: 5px 10px;
                border-radius: 999px;
                font-weight: 900;
            }

            .iq-download {
                border-radius: 22px;
                padding: 22px;
                background: linear-gradient(135deg, #0f172a, #1d4ed8);
                color: #ffffff;
                box-shadow: 0 16px 42px rgba(37,99,235,.22);
                margin-top: 18px;
            }

            .iq-download h3 {
                color: white;
                margin: 0 0 6px 0;
            }

            .iq-download p {
                color: #dbeafe;
                margin: 0;
            }

            .iq-agent-box {
                border: 1px solid #bfdbfe;
                background: #eff6ff;
                border-radius: 20px;
                padding: 20px;
                color: #0f172a;
                font-size: 1.02rem;
                line-height: 1.55;
            }

            .stButton > button {
                border-radius: 14px;
                font-weight: 800;
                border: 0;
                padding: .7rem 1.1rem;
            }

            div[data-testid="stDownloadButton"] button {
                border-radius: 14px;
                font-weight: 900;
                background: #16a34a;
                color: white;
                border: 0;
            }

            div[data-testid="stFileUploader"] {
                border: 1px dashed #93c5fd;
                border-radius: 18px;
                padding: 10px;
                background: rgba(239,246,255,.55);
            }

            .stTabs [data-baseweb="tab-list"] {
                gap: 10px;
            }

            .stTabs [data-baseweb="tab"] {
                border-radius: 999px;
                padding: 10px 16px;
                background: #f1f5f9;
                font-weight: 800;
            }
        </style>
        """,
        unsafe_allow_html=True,
    )


def render_hero() -> None:
    st.markdown(
        """
        <div class="iq-hero">
            <div class="iq-eyebrow">⚡ POLISHED UI v2 • Hackathon Demo • Synthetic Invoices Only</div>
            <div class="iq-title">Invoice IQ Agent</div>
            <div class="iq-subtitle">
                PDF energy invoices → extracted fields → validation rules → explainable exceptions → clean Excel workbook.
            </div>
            <div class="iq-chip-row">
                <span class="iq-chip">📄 Multi-PDF upload</span>
                <span class="iq-chip">🧠 Microsoft Foundry Agent Q&amp;A</span>
                <span class="iq-chip">✅ Rule validation</span>
                <span class="iq-chip">📊 Review dashboard</span>
                <span class="iq-chip">📥 Excel export</span>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_flow() -> None:
    steps = [
        ("Upload", "Select the 15 synthetic PDF invoices."),
        ("Extract", "Read the requested electricity-invoice fields."),
        ("Validate", "Check required fields, Bill Total, 4CP, and line items."),
        ("Explain", "Ask the agent about totals, flags, columns, and spend."),
        ("Export", "Download Summary, Line Items, Flags, and Validation tabs."),
    ]
    cols = st.columns(5)
    for idx, (title, desc) in enumerate(steps, start=1):
        with cols[idx - 1]:
            st.markdown(
                f"""
                <div class="iq-flow">
                    <div class="iq-step-number">{idx}</div>
                    <h4>{title}</h4>
                    <p>{desc}</p>
                </div>
                """,
                unsafe_allow_html=True,
            )


def render_kpi(label: str, value: str, caption: str) -> None:
    st.markdown(
        f"""
        <div class="iq-kpi">
            <div class="iq-kpi-label">{label}</div>
            <div class="iq-kpi-value">{value}</div>
            <div class="iq-kpi-caption">{caption}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def style_validation_table(df: pd.DataFrame):
    status_cols = [
        col
        for col in ["Review Status", "Required Fields", "Bill Total Math", "4CP Math", "Line Item Sum"]
        if col in df.columns
    ]

    def color_status(value: Any) -> str:
        text = str(value).upper()
        if text == "PASS":
            return "background-color: #ecfdf5; color: #047857; font-weight: 800;"
        if text in {"REVIEW", "FAIL"}:
            return "background-color: #fef2f2; color: #b91c1c; font-weight: 800;"
        if text == "WARN":
            return "background-color: #fffbeb; color: #b45309; font-weight: 800;"
        return ""

    if not status_cols:
        return df
    return df.style.map(color_status, subset=status_cols)


def render_landing_cards() -> None:
    left, mid, right = st.columns(3)
    with left:
        st.markdown(
            """
            <div class="iq-card">
                <h3>🎯 Demo objective</h3>
                <p>Show a complete agent workflow: extraction, validation, exception handling, explanation, and Excel output.</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with mid:
        st.markdown(
            """
            <div class="iq-card">
                <h3>📁 Best demo input</h3>
                <p>Upload all 15 PDFs from the <b>sample_invoices</b> folder. Invoices 14 and 15 are intentionally flagged.</p>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with right:
        st.markdown(
            """
            <div class="iq-card">
                <h3>🏁 Judging value</h3>
                <p>The app is visual, auditable, and business-facing. It makes exceptions clear instead of hiding them.</p>
            </div>
            """,
            unsafe_allow_html=True,
        )


def render_charts(summary_df: pd.DataFrame, flags_df: pd.DataFrame, validation_df: pd.DataFrame) -> None:
    spend_df = summary_df.copy()
    spend_df["Bill Total"] = pd.to_numeric(spend_df["Bill Total"], errors="coerce")
    spend_df["Usage - Actual KWH"] = pd.to_numeric(spend_df["Usage - Actual KWH"], errors="coerce")
    spend_df["Production Month"] = pd.to_datetime(spend_df["Production Month"], errors="coerce")

    chart_left, chart_right = st.columns([1.25, 1])

    with chart_left:
        st.markdown("### 📈 Spend by production month")
        monthly = (
            spend_df.dropna(subset=["Production Month"])
            .assign(Month=lambda d: d["Production Month"].dt.strftime("%b %Y"))
            .groupby("Month", sort=False)["Bill Total"]
            .sum()
            .reset_index()
        )
        if monthly.empty:
            st.info("No production month data available for charting.")
        else:
            st.bar_chart(monthly.set_index("Month"), use_container_width=True)

    with chart_right:
        st.markdown("### 🧾 Review status")
        if validation_df.empty:
            st.info("No validation data available.")
        else:
            status_counts = validation_df["Review Status"].value_counts().rename_axis("Status").reset_index(name="Count")
            st.bar_chart(status_counts.set_index("Status"), use_container_width=True)

    chart_left, chart_right = st.columns([1.25, 1])

    with chart_left:
        st.markdown("### 🔎 Top invoices by Bill Total")
        top_spend = spend_df[["Source File", "Bill Total"]].dropna().sort_values("Bill Total", ascending=False).head(8)
        if top_spend.empty:
            st.info("No Bill Total values available.")
        else:
            st.bar_chart(top_spend.set_index("Source File"), use_container_width=True)

    with chart_right:
        st.markdown("### 🚩 Flag volume by rule")
        if flags_df.empty:
            st.success("No flags generated.")
        else:
            flag_counts = flags_df["Rule"].value_counts().rename_axis("Rule").reset_index(name="Count")
            st.bar_chart(flag_counts.set_index("Rule"), use_container_width=True)


def main() -> None:
    st.set_page_config(
        page_title=APP_TITLE,
        page_icon="⚡",
        layout="wide",
        initial_sidebar_state="expanded",
    )
    inject_professional_css()

    with st.sidebar:
        st.markdown("# ⚡ Invoice IQ")
        st.markdown("**Polished demo control panel**")
        st.markdown("---")
        st.markdown("### 1. Upload sample PDFs")
        uploaded_files = st.file_uploader(
            "Upload invoice PDFs",
            type=["pdf"],
            accept_multiple_files=True,
            help="Use Ctrl+A inside the sample_invoices folder to upload all 15 PDFs.",
        )
        st.caption("Recommended: upload all 15 synthetic invoices.")

        process_clicked = st.button(
            "🚀 Run AI extraction",
            type="primary",
            use_container_width=True,
            disabled=not uploaded_files,
        )

        st.markdown("---")
        st.markdown("### 2. Demo checklist")
        st.markdown(
            """
            ✅ Extract requested columns  
            ✅ Validate bill math  
            ✅ Flag exceptions  
            ✅ Ask the agent  
            ✅ Export Excel workbook
            """
        )
        st.markdown("---")
        st.caption("Synthetic/sample invoices only — no confidential data.")
        st.caption("Foundry endpoint comes from .env → AGENT_ENDPOINT.")

    render_hero()
    render_flow()
    st.markdown("<br>", unsafe_allow_html=True)

    if process_clicked:
        with st.spinner("Processing PDFs, extracting fields, validating rules, and preparing dashboard..."):
            summary_df, line_items_df, flags_df, validation_df, raw_text = process_uploaded_files(uploaded_files)
            st.session_state["summary_df"] = summary_df
            st.session_state["line_items_df"] = line_items_df
            st.session_state["flags_df"] = flags_df
            st.session_state["validation_df"] = validation_df
            st.session_state["raw_text"] = raw_text
            st.session_state["uploaded_names"] = [file.name for file in uploaded_files]
        st.toast("Invoice IQ processing complete.", icon="✅")

    if "summary_df" not in st.session_state:
        render_landing_cards()
        st.info("Upload the sample PDFs from the sidebar, then click **Run AI extraction**.")
        return

    summary_df = st.session_state["summary_df"]
    line_items_df = st.session_state["line_items_df"]
    flags_df = st.session_state["flags_df"]
    validation_df = st.session_state["validation_df"]

    processed_count = len(summary_df)
    review_count = int((validation_df["Review Status"] == "REVIEW").sum()) if not validation_df.empty else 0
    pass_count = processed_count - review_count
    total_bill = pd.to_numeric(summary_df["Bill Total"], errors="coerce").sum()
    total_usage = pd.to_numeric(summary_df["Usage - Actual KWH"], errors="coerce").sum()
    avg_rate = total_bill / total_usage if total_usage else 0
    avg_confidence = pd.to_numeric(validation_df["Confidence"], errors="coerce").mean() if not validation_df.empty else 0

    k1, k2, k3, k4, k5 = st.columns(5)
    with k1:
        render_kpi("Invoices", f"{processed_count}", "PDFs processed")
    with k2:
        render_kpi("Passed", f"{pass_count}", "No critical exceptions")
    with k3:
        render_kpi("Needs review", f"{review_count}", "Flagged for analyst review")
    with k4:
        render_kpi("Total spend", f"${total_bill:,.0f}", "Combined Bill Total")
    with k5:
        render_kpi("Avg $/KWH", f"${avg_rate:.4f}", f"Confidence {avg_confidence:.0%}")

    st.markdown("<br>", unsafe_allow_html=True)
    render_charts(summary_df, flags_df, validation_df)

    st.markdown("---")

    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
        "📋 Summary",
        "🧾 Line Items",
        "🚩 Flags / Exceptions",
        "✅ Validation",
        "🤖 Ask the Foundry Agent",
        "🧪 Raw PDF Text",
    ])

    with tab1:
        st.markdown("### 📋 Summary — requested electricity-invoice output columns")
        st.caption("This tab is designed to match the final Excel Summary output.")
        st.dataframe(clean_display_df(summary_df[OUTPUT_COLUMNS]), use_container_width=True, height=470)

    with tab2:
        st.markdown("### 🧾 Extracted line items")
        if line_items_df.empty:
            st.warning("No line items were extracted.")
        else:
            st.dataframe(line_items_df, use_container_width=True, height=470)

    with tab3:
        st.markdown("### 🚩 Flags / Exceptions")
        if flags_df.empty:
            st.success("No flags or exceptions. All invoices passed the current demo rules.")
        else:
            critical = int((flags_df["Severity"] == "Critical").sum())
            warning = int((flags_df["Severity"] == "Warning").sum())
            c1, c2 = st.columns(2)
            c1.metric("Critical flags", critical)
            c2.metric("Warnings", warning)
            st.dataframe(flags_df, use_container_width=True, height=470)

    with tab4:
        st.markdown("### ✅ Validation dashboard")
        st.dataframe(style_validation_table(validation_df), use_container_width=True, height=470)

    with tab5:
        st.markdown("### 🤖 Ask the Foundry Agent")
        st.caption("Try: Do totals match? Which invoices were flagged? What is the average $/KWH? What columns were extracted?")
        quick_prompts = st.columns(4)
        default_question = ""
        if quick_prompts[0].button("Do totals match?", use_container_width=True):
            default_question = "Do totals match?"
        if quick_prompts[1].button("Which invoices are flagged?", use_container_width=True):
            default_question = "Which invoices are flagged?"
        if quick_prompts[2].button("Average $/KWH", use_container_width=True):
            default_question = "What is the average cost per KWH?"
        if quick_prompts[3].button("Extracted columns", use_container_width=True):
            default_question = "What columns were extracted?"

        question = st.text_input("Ask a question", value=default_question, placeholder="Do totals match?")
        if question:
            answer = ask_agent(question, summary_df, line_items_df, flags_df, validation_df)
            st.markdown(
                f"""
                <div class="iq-agent-box">
                    <b>Agent:</b> {answer}
                </div>
                """,
                unsafe_allow_html=True,
            )

    with tab6:
        st.markdown("### 🧪 Extracted PDF text")
        st.text_area("Raw extracted text", st.session_state["raw_text"], height=430)

    excel_bytes = build_excel_workbook(summary_df, line_items_df, flags_df, validation_df)
    st.markdown(
        """
        <div class="iq-download">
            <h3>📥 Export clean workbook</h3>
            <p>Workbook includes Summary, Line Items, Flags_Exceptions, and Validation sheets.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.download_button(
        label="Download clean Excel workbook",
        data=excel_bytes,
        file_name="invoice_iq_energy_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


if __name__ == "__main__":
    main()
